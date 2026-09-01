import io
import json
import logging
import os
import re
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.staticfiles import finders
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import load_workbook

from .forms import BusForm, LoginForm, SchoolRegisterForm, StudentForm

User = get_user_model()
from .models import (
    Absence,
    Bus,
    BusMaintenance,
    GPSLog,
    Incident,
    Route,
    School,
    Student,
)
from .services import (
    DELAY_THRESHOLD_MINUTES,
    SPEED_LIMIT_KMH,
    absent_student_ids,
    estimate_bus_delay,
    estimate_remaining_minutes,
    haversine_km,
    recalculate_school_routes,
    send_parent_sms,
)
from .signals import recalc_hold

logger = logging.getLogger(__name__)

STUDENT_IMPORT_COLUMNS = ["matricule", "nom", "postnom", "prenom", "adresse", "telephone"]


def _unassigned_students_with_gps(school):
    """Élèves actifs ayant une position GPS valide mais aucun bus affecté
    (inclut les élèves en surcapacité)."""
    return Student.objects.filter(
        school=school,
        is_active=True,
        latitude__isnull=False,
        longitude__isnull=False,
        assigned_route__isnull=True,
    )
BUS_IMPORT_COLUMNS = ["code_bus", "nom_chauffeur", "capacite"]


def _bus_statistics(school=None):
    """Par bus (de l'école donnée, ou tous) : élèves assignés, pris, restants."""
    stats = []
    buses = Bus.objects.filter(school=school).order_by("code_bus") if school else Bus.objects.all().order_by("code_bus")
    for bus in buses:
        routes = bus.routes.all()
        assigned = sum(route.students.count() for route in routes)
        taken = sum(route.students_taken for route in routes)
        remaining = sum(route.students_remaining for route in routes)
        pickup_rate = round(taken / assigned * 100, 1) if assigned else 0.0
        stats.append(
            {
                "bus": bus,
                "routes": routes,
                "assigned": assigned,
                "taken": taken,
                "remaining": remaining,
                "pickup_rate": pickup_rate,
            }
        )
    return stats


def _student_for_user(user):
    """Élève associé à l'utilisateur : compte élève (Student.user) ou parent (parent_user)."""
    student = getattr(user, "student", None)
    if student is not None:
        return student
    return user.children.order_by("id").first()


def _pending_password_change(user):
    """Profil (Student ou Bus) exigeant un changement de mot de passe, sinon None."""
    student = getattr(user, "student", None)
    if student is not None and student.must_change_password:
        return student
    bus = getattr(user, "driven_bus", None)
    if bus is not None and bus.must_change_password:
        return bus
    return None


def _user_school(user):
    """École de l'utilisateur : son école si admin d'école, sinon la première (démo)."""
    school = getattr(user, "school", None)
    return school or School.objects.first()


def _school_scope_guard(request):
    """Garde des vues rattachées à une école :
    - l'administration générale (superuser, sans école) est redirigée vers la
      gestion des écoles (/ecoles/) ;
    - une école désactivée est déconnectée et renvoyée vers /login/.
    Retourne une HttpResponse de redirection, ou None si l'accès est autorisé."""
    if request.user.is_superuser:
        return redirect("routing:schools_admin")
    school = _user_school(request.user)
    if school is not None and not school.is_active:
        messages.error(request, "Votre école a été désactivée par l'administration générale.")
        logout(request)
        return redirect("routing:login")
    return None


def _ensure_student_account(student, password=None):
    """Crée/récupère le compte de connexion de l'élève (identifiant = matricule)."""
    if student.user_id:
        return student.user
    user, created = User.objects.get_or_create(username=student.matricule)
    if created or not user.has_usable_password():
        user.set_password(password or student.matricule)
        user.save()
    student.user = user
    student.save(update_fields=["user"])
    return user


def _ensure_driver_account(bus):
    """Crée/récupère le compte de connexion du chauffeur d'un bus
    (identifiant = code du bus, mot de passe = code du bus au départ)."""
    if bus.driver_user_id:
        return bus.driver_user
    # Un code de bus peut exister en double (saisie) : le compte étant lié en
    # OneToOne à un seul bus, on suffixe le nom d'utilisateur dans ce cas.
    username = bus.code_bus
    user, created = User.objects.get_or_create(username=username)
    if not created and hasattr(user, "driven_bus") and user.driven_bus.pk != bus.pk:
        i = 2
        while User.objects.filter(username=f"{username}_{i}").exists():
            i += 1
        username = f"{username}_{i}"
        user, created = User.objects.get_or_create(username=username)
    if created or not user.has_usable_password():
        user.set_password(bus.code_bus)
        user.save()
    bus.driver_user = user
    bus.save(update_fields=["driver_user"])
    return user


# ---------------------------------------------------------------- PWA (hors-ligne)


# ---------------------------------------------------------------- SEO : robots.txt & sitemap.xml


def robots_txt_view(request):
    """robots.txt — indique aux moteurs de recherche quelles pages indexer."""
    from django.conf import settings
    content = (
        f"User-agent: *\n"
        f"Allow: /\n"
        f"Disallow: /admin/\n"
        f"Disallow: /dashboard/\n"
        f"Disallow: /driver/\n"
        f"Disallow: /parent/\n"
        f"Disallow: /login/\n"
        f"Disallow: /logout/\n"
        f"Disallow: /import/\n"
        f"Disallow: /export/\n"
        f"Disallow: /api/\n"
        f"Disallow: /sw.js\n"
        f"\nSitemap: {settings.SITE_URL}/sitemap.xml\n"
    )
    return HttpResponse(content, content_type="text/plain; charset=utf-8")


def sitemap_xml_view(request):
    """sitemap.xml — toutes les pages publiques indexables par Google."""
    from django.conf import settings
    from django.urls import reverse

    base = settings.SITE_URL
    pages = [
        {
            'loc': f"{base}/",
            'priority': '1.0',
            'changefreq': 'daily',
        },
        {
            'loc': f"{base}{reverse('routing:about')}",
            'priority': '0.8',
            'changefreq': 'monthly',
        },
        {
            'loc': f"{base}{reverse('routing:login')}",
            'priority': '0.5',
            'changefreq': 'monthly',
        },
        {
            'loc': f"{base}{reverse('routing:register_school')}",
            'priority': '0.6',
            'changefreq': 'monthly',
        },
        {
            'loc': f"{base}{reverse('routing:manual')}",
            'priority': '0.7',
            'changefreq': 'monthly',
        },
    ]

    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for page in pages:
        xml += '  <url>\n'
        xml += f'    <loc>{page["loc"]}</loc>\n'
        xml += f'    <changefreq>{page["changefreq"]}</changefreq>\n'
        xml += f'    <priority>{page["priority"]}</priority>\n'
        xml += '  </url>\n'
    xml += '</urlset>'
    return HttpResponse(xml, content_type='application/xml; charset=utf-8')


def pwa_service_worker(request):
    """Sert le Service Worker à la racine (/sw.js).

    Le fichier source vit dans routing/static/js/sw.js ; il est resservi depuis
    la racine avec l'en-tête `Service-Worker-Allowed: /` afin que sa portée
    couvre tout le site (accueil, chauffeur, parent, école).
    """
    path = finders.find("js/sw.js")
    if path is None:
        return HttpResponse("Service Worker introuvable.", status=404)
    with open(path, "rb") as f:
        content = f.read()
    response = HttpResponse(content, content_type="application/javascript; charset=utf-8")
    response["Service-Worker-Allowed"] = "/"
    # Jamais de cache agressif : les mises à jour doivent se propager rapidement.
    response["Cache-Control"] = "no-cache"
    return response


def _seo_context(request, title=None, description=None, url_path='/'):
    """Contexte commun SEO pour toutes les pages publiques."""
    from django.conf import settings
    return {
        'SITE_NAME': settings.SITE_NAME,
        'SITE_URL': settings.SITE_URL,
        'SITE_DESCRIPTION': settings.SITE_DESCRIPTION,
        'SITE_KEYWORDS': settings.SITE_KEYWORDS,
        'GOOGLE_SITE_VERIFICATION': settings.GOOGLE_SITE_VERIFICATION,
        'page_title': title or settings.SITE_NAME,
        'page_description': description or settings.SITE_DESCRIPTION,
        'page_url': f"{settings.SITE_URL}{url_path}",
    }


def home_view(request):
    """Page d'accueil publique du système de transport scolaire."""
    ctx = _seo_context(
        request,
        title='Scolaloop — Système de Transport Scolaire Intelligent | ESTECH',
        description='Scolaloop optimise les trajets de bus scolaires en temps réel : tournées en boucle fermée, itinéraires routiers OSRM, suivi GPS et gestion des élèves. Développé par ESTECH.',
        url_path='/',
    )
    return render(request, "routing/home.html", ctx)


def about_view(request):
    """Page « À propos » : présentation du projet et de l'entreprise ESTECH."""
    ctx = _seo_context(
        request,
        title='À propos — Scolaloop · ESTECH',
        description='Découvrez ScolaLoop, le système intelligent de transport scolaire développé par ESTECH : optimisation des tournées, suivi GPS en temps réel et sécurité routière.',
        url_path='/apropos/',
    )
    return render(request, "routing/about.html", ctx)


def manual_view(request):
    """Manuel d'utilisation public (sans connexion) de l'application ScolaLoop."""
    ctx = _seo_context(
        request,
        title='Manuel d\'utilisation — Scolaloop · ESTECH',
        description='Guide complet d\'utilisation de ScolaLoop : transport scolaire, suivi GPS, gestion des élèves, tournées et interface parent.',
        url_path='/manuel/',
    )
    return render(request, "routing/manual.html", ctx)


def manual_pdf_view(request):
    """Génère et télécharge le manuel d'utilisation au format PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, ListFlowable, ListItem,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
        title="Manuel d'utilisation ScolaLoop",
        author="ESTECH",
    )

    styles = getSampleStyleSheet()
    indigo = HexColor="#4F46E5"
    blue = HexColor("#2563EB")
    slate = HexColor("#0f172a")
    muted = HexColor("#475569")

    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=26,
                                  textColor=indigo, spaceAfter=12, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11,
                                     textColor=muted, alignment=TA_CENTER, spaceAfter=30)
    h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=18, textColor=slate,
                         spaceBefore=20, spaceAfter=10, borderPadding=(0, 0, 4, 0))
    h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=14, textColor=indigo,
                         spaceBefore=14, spaceAfter=8)
    body = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=15,
                           textColor=muted, alignment=TA_JUSTIFY, spaceAfter=6)
    bullet = ParagraphStyle('Bullet', parent=body, leftIndent=18, bulletIndent=6,
                             spaceBefore=2, spaceAfter=2)
    tip = ParagraphStyle('Tip', parent=body, leftIndent=12, backColor=HexColor("#f0fdf4"),
                          borderPadding=(6, 8, 6, 8), spaceBefore=6, spaceAfter=8)

    story = []

    # --- Page de garde ---
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph(" Manuel d'utilisation", title_style))
    story.append(Paragraph("Scolaloop - Système de Transport Scolaire Intelligent", subtitle_style))
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph("Version 1.0 - 2026", subtitle_style))
    story.append(Paragraph("Développé par ESTECH", subtitle_style))
    story.append(PageBreak())

    # --- Sommaire ---
    story.append(Paragraph("Sommaire", h1))
    toc = [
        "1. Présentation générale",
        "2. Connexion et sécurité",
        "3. Tableau de bord (Administration)",
        "4. Gestion des élèves",
        "5. Gestion de la flotte de bus",
        "6. Importation Excel",
        "7. Optimisation des tournées",
        "8. Suivi GPS en temps réel",
        "9. Interface chauffeur",
        "10. Interface parent / élève",
        "11. Statistiques",
        "12. Maintenance de la flotte",
        "13. Administration générale",
    ]
    for item in toc:
        story.append(Paragraph(item, body))
    story.append(PageBreak())

    # --- Contenu ---
    sections = [
        ("1. Présentation générale", [
            ("<b>ScolaLoop</b> est un système intelligent de transport scolaire qui optimise les trajets de bus scolaires en temps réel. Il combine :", []),
            ("", [
                "• <b>Optimisation des tournées</b> - Problème VRP résolu par Google OR-Tools",
                "• <b>Itinéraires routiers réels</b> - Calcul via OSRM (boucle École - élèves - École)",
                "• <b>Suivi GPS en temps réel</b> - Position des bus, alertes, replay",
                "• <b>Communication parents-école</b> - Absences, SMS d'urgence, interface dédiée",
                "• <b>Gestion des élèves</b> - Inscription, affectation automatique, import Excel",
            ]),
            ("Le système s'adresse à trois types d'utilisateurs : <b>l'administration de l'école</b>, <b>les chauffeurs</b> et <b>les parents / élèves</b>.", []),
        ]),
        ("2. Connexion et sécurité", [
            ("Chaque utilisateur dispose d'un compte sécurisé :", []),
            ("", [
                "• <b>École (admin)</b> : identifiant = code de l'école, mot de passe défini à l'inscription",
                "• <b>Chauffeur</b> : identifiant = code du bus, mot de passe = code du bus (à changer à la 1ère connexion)",
                "• <b>Parent / Élève</b> : identifiant = matricule, mot de passe = matricule (à changer à la 1ère connexion)",
            ]),
            ("À la première connexion, un changement de mot de passe est <b>obligatoire</b>. L'interface parent nécessite ensuite de positionner le domicile sur une carte.", []),
            ("Conseil : les mots de passe par défaut (matricule ou code bus) doivent être remplacés dès la première connexion pour garantir la sécurité.", []),
        ]),
        ("3. Tableau de bord (Administration)", [
            ("Le tableau de bord est l'écran principal de l'administration de l'école. Il affiche :", []),
            ("", [
                "• <b>Carte interactive</b> - Positions des bus, élèves et arrêts sur Leaflet",
                "• <b>Liste des bus</b> - Code, chauffeur, capacité, statut",
                "• <b>Liste des élèves</b> - Matricule, nom, bus assigné",
                "• <b>Alertes en direct</b> - Retards, excès de vitesse, incidents",
                "• <b>Statistiques</b> - Taux de ramassage, élèves pris/restants par bus",
            ]),
            ("Le panneau latéral permet d'accéder rapidement à toutes les fonctionnalités : ajout d'élèves, import Excel, maintenance, etc.", []),
        ]),
        ("4. Gestion des élèves", [
            ("<b>Ajouter un élève</b> :", []),
            ("", [
                "• Remplir le formulaire (matricule, nom, postnom, prénom, adresse, téléphone parent)",
                "• L'affectation au bus est <b>automatique</b> (signal post_save, moteur VRP)",
                "• Un compte de connexion est créé automatiquement (identifiant = matricule)",
            ]),
            ("<b>Modifier / Supprimer</b> : cliquer sur l'icône correspondante dans la liste.", []),
            ("<b>Geler un élève</b> : exclut temporairement l'élève des tournées sans le supprimer.", []),
            ("<b>Réinitialiser le mot de passe</b> : l'élève devra le personnaliser à sa prochaine connexion.", []),
            ("Import Excel : pour inscrire plusieurs élèves d'un coup, utilisez l'import Excel (colonnes : matricule, nom, postnom, prénom, adresse, téléphone).", []),
        ]),
        ("5. Gestion de la flotte de bus", [
            ("<b>Ajouter un bus</b> :", []),
            ("", [
                "• Remplir le formulaire (code bus, nom du chauffeur, capacité)",
                "• Un compte chauffeur est créé automatiquement (identifiant = code bus)",
            ]),
            ("<b>Modifier / Supprimer</b> : opérations classiques depuis la liste.", []),
            ("<b>Réinitialiser le mot de passe chauffeur</b> : le chauffeur devra le personnaliser.", []),
            ("Import Excel : pour ajouter plusieurs bus, utilisez l'import Excel (colonnes : code_bus, nom_chauffeur, capacite).", []),
        ]),
        ("6. Importation Excel", [
            ("ScolaLoop supporte l'import en masse via des fichiers Excel (.xlsx) :", []),
            ("", [
                "• <b>Élèves</b> : matricule, nom, postnom, prénom, adresse, téléphone",
                "• <b>Bus</b> : code_bus, nom_chauffeur, capacite",
                "• <b>Élèves (gel)</b> : même format, mais importe uniquement les statuts de gel",
            ]),
            ("L'import vérifie automatiquement les doublons et valide les données avant insertion.", []),
        ]),
        ("7. Optimisation des tournées", [
            ("Le moteur d'optimisation fonctionne automatiquement :", []),
            ("", [
                "• À chaque ajout/modification d'élève, le signal post_save relance le calcul",
                "• L'algorithme VRP (Vehicle Routing Problem) répartit les élèves sur les bus",
                "• L'ordre des arrêts est calculé par OR-Tools (TSP interne par bus)",
                "• Les itinéraires routiers réels sont obtenus via OSRM",
                "• Le résultat est une boucle : École - arrêts - École",
            ]),
            ("Recalcul : si vous modifiez la position d'un élève, les tournées sont recalculées automatiquement.", []),
        ]),
        ("8. Suivi GPS en temps réel", [
            ("Le suivi GPS permet de :", []),
            ("", [
                "• <b>Localiser les bus</b> en temps réel sur la carte",
                "• <b>Recevoir des alertes</b> d'excès de vitesse (au-delà de la limite configurée)",
                "• <b>Estimer les retards</b> par rapport à l'horaire prévu",
                "• <b>Rejouer les trajets</b> passés via le replay GPS",
            ]),
            ("La position GPS est envoyée par le chauffeur via l'application mobile ou le formulaire web dédié.", []),
        ]),
        ("9. Interface chauffeur", [
            ("Le chauffeur dispose d'une interface simplifiée :", []),
            ("", [
                "• <b>Carte du trajet</b> - L'itinéraire à suivre avec les arrêts",
                "• <b>Marquage des élèves</b> - Confirmer la prise en charge de chaque élève",
                "• <b>Envoi de position</b> - Mettre à jour la position GPS",
                "• <b>Signalement d'incident</b> - Déclarer un problème (panne, accident, etc.)",
                "• <b>Synchronisation hors-ligne</b> - Les données accumulées sans connexion sont envoyées automatiquement",
            ]),
        ]),
        ("10. Interface parent / élève", [
            ("L'interface parent permet de :", []),
            ("", [
                "• <b>Suivre le bus</b> en temps réel sur la carte",
                "• <b>Voir l'état du ramassage</b> : statut du trajet, estimation du temps d'arrivée",
                "• <b>Signaler une absence</b> - Le bus ne s'arrêtera pas chez vous",
                "• <b>Afficher son domicile</b> sur la carte",
            ]),
            ("Au premier accès, le parent doit positionner le domicile de l'élève sur la carte (clic ou position GPS).", []),
            ("Notification : une alerte est envoyée lorsque le bus est proche du domicile (100 m).", []),
        ]),
        ("11. Statistiques", [
            ("La page statistiques fournit un tableau d'ensemble :", []),
            ("", [
                "• Nombre d'élèves assignés / pris / restants par bus",
                "• Taux de ramassage par bus",
                "• Distance totale et durée estimée des tournées",
                "• Évolution dans le temps",
            ]),
        ]),
        ("12. Maintenance de la flotte", [
            ("L'administration peut enregistrer les interventions d'entretien :", []),
            ("", [
                "• <b>Type d'intervention</b> : révision, changement de pneus, vidange, etc.",
                "• <b>Coût</b> et <b>date</b> de l'intervention",
                "• <b>Prochaine échéance</b> (kilométrage ou date)",
            ]),
            ("Un historique complet par bus est disponible depuis le tableau de bord.", []),
        ]),
        ("13. Administration générale", [
            ("L'administration générale (superutilisateur) gère l'ensemble des écoles :", []),
            ("", [
                "• <b>Liste des écoles</b> - Activer / Désactiver",
                "• <b>Statistiques globales</b> - Nombre total d'élèves, bus, routes",
                "• <b>Création d'écoles</b> - Inscription directe",
                "• <b>Suppression</b> - Suppression définitive avec toutes les données",
                "• <b>Réinitialisation des mots de passe</b> des comptes école",
            ]),
        ]),
    ]

    for section_title, paragraphs in sections:
        story.append(Paragraph(section_title, h1))
        for text, bullets in paragraphs:
            if text:
                story.append(Paragraph(text, body))
            for b in bullets:
                story.append(Paragraph(b, bullet))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="manuel_scolaloop.pdf"'
    return response


def login_view(request):
    """Connexion par matricule + mot de passe."""
    if request.user.is_authenticated:
        return redirect("routing:dashboard")

    form = LoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = authenticate(
            request,
            username=form.cleaned_data["matricule"],
            password=form.cleaned_data["password"],
        )
        if user is not None:
            # Une école désactivée par l'administration ne peut plus se connecter.
            school = getattr(user, "school", None)
            if school is not None and not school.is_active:
                messages.error(
                    request,
                    "Cette école a été désactivée par l'administration générale. "
                    "Contactez l'administrateur.",
                )
                return render(request, "routing/login.html", {"form": form})
            login(request, user)
            if _pending_password_change(user):
                return redirect("routing:change_password")
            if user.is_superuser:  # administration générale → gestion des écoles
                return redirect("routing:schools_admin")
            if hasattr(user, "driven_bus"):  # chauffeur → interface dédiée
                # Présence en direct : le bus apparaît sur la carte de l'école.
                driven_bus = user.driven_bus
                if not driven_bus.driver_connected:
                    driven_bus.driver_connected = True
                    driven_bus.save(update_fields=["driver_connected"])
                return redirect("routing:driver")
            student = _student_for_user(user)
            if student is not None:  # élève/parent → interface parent
                if student.latitude is None or student.longitude is None:
                    return redirect("routing:parent_setup_home")
                return redirect("routing:parent")
            return redirect("routing:dashboard")
        messages.error(request, "Matricule ou mot de passe incorrect.")
    return render(request, "routing/login.html", {"form": form})


def register_school_view(request):
    """Inscription d'une nouvelle école (code, nom, adresse, coordonnées, mot de passe).

    Accessible à tous (création sur place) ; l'administration générale (superuser)
    peut aussi créer une école directement.
    """
    if request.user.is_authenticated and not request.user.is_superuser:
        return redirect("routing:dashboard")

    ctx = _seo_context(
        request,
        title='Créer une École — Scolaloop | Inscription Transport Scolaire',
        description='Inscrivez votre école sur Scolaloop et commencez à optimiser vos trajets de bus scolaires.',
        url_path='/register-school/',
    )

    form = SchoolRegisterForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        school = form.save()
        user = User.objects.create_user(
            username=school.code_ecole,
            password=form.cleaned_data["password"],
        )
        school.user = user
        school.save(update_fields=["user"])
        # Backend explicite : plusieurs backends sont configurés.
        login(request, user, backend="django.contrib.auth.backends.ModelBackend")
        messages.success(
            request,
            f"École « {school.name} » créée ({school.code_ecole}). "
            "Elle démarre avec 0 élève et 0 bus.",
        )
        if request.user.is_superuser:
            return redirect("routing:schools_admin")
        return redirect("routing:dashboard")
    return render(request, "routing/register_school.html", {"form": form, **ctx})


@login_required
def change_password_view(request):
    """Changement de mot de passe obligatoire à la première connexion."""
    profile = _pending_password_change(request.user)
    if profile is None:
        # Rien à changer : direction l'interface correspondante.
        if hasattr(request.user, "driven_bus"):
            return redirect("routing:driver")
        if _student_for_user(request.user) is not None:
            return redirect("routing:parent")
        return redirect("routing:dashboard")

    form = SetPasswordForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        # Le mot de passe a changé : le hash de session doit suivre, sinon
        # l'utilisateur serait déconnecté au prochain accès.
        update_session_auth_hash(request, request.user)
        profile.must_change_password = False
        profile.save(update_fields=["must_change_password"])
        messages.success(request, "Mot de passe mis à jour avec succès.")
        if hasattr(request.user, "driven_bus"):
            return redirect("routing:driver")
        student = _student_for_user(request.user)
        if student is not None:
            # Étape 2 de l'onboarding : position du domicile sur la carte.
            if student.latitude is None or student.longitude is None:
                return redirect("routing:parent_setup_home")
            return redirect("routing:parent")
        return redirect("routing:dashboard")
    return render(request, "routing/change_password.html", {"form": form})


def logout_view(request):
    """Déconnexion puis retour à l'accueil.

    Si l'utilisateur est un chauffeur, son bus disparaît immédiatement
    de la carte de l'école (driver_connected=False).
    """
    driven_bus = getattr(request.user, "driven_bus", None)
    if driven_bus is not None and driven_bus.driver_connected:
        driven_bus.driver_connected = False
        driven_bus.save(update_fields=["driver_connected"])
    logout(request)
    return redirect("routing:home")


@login_required
def dashboard_view(request):
    """Tableau de bord (protégé) : carte en boucle fermée, élèves, flotte, statistiques.

    Réservé à l'école / l'administration : un chauffeur est redirigé vers /driver/,
    un élève ou parent vers /parent/.
    """
    if _pending_password_change(request.user):
        return redirect("routing:change_password")
    if hasattr(request.user, "driven_bus"):
        return redirect("routing:driver")
    if _student_for_user(request.user) is not None:
        return redirect("routing:parent")
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked

    school = _user_school(request.user)
    students = Student.objects.filter(school=school, is_active=True).order_by("nom", "prenom")
    buses = Bus.objects.filter(school=school).order_by("code_bus")
    routes = (
        Route.objects.filter(school=school)
        .prefetch_related("stops__stop", "bus", "school")
        .order_by("-created_at")
    )

    # Marqueurs carte : tous les bus actifs apparaissent sur la carte.
    # Chauffeur connecte -> position GPS en direct ; sinon -> position de base (ecole).
    bus_markers = []
    delayed_buses = 0
    speed_alerts = 0
    now = timezone.now()
    for offset, b in enumerate(buses):
        if not b.is_in_service:
            continue
        stale = (
            b.last_position_at is not None
            and (now - b.last_position_at).total_seconds() > 300
        )
        route = b.routes.order_by("-created_at").first()
        if b.last_latitude is not None and b.last_longitude is not None:
            lat, lon = b.last_latitude, b.last_longitude
            position_label = (
                f"Dernière position GPS : {b.last_position_at.strftime('%H:%M:%S')}"
                if b.last_position_at else "Position GPS enregistrée"
            )
        else:
            lat = (school.latitude + offset * 0.0012) if school else 0.0
            lon = (school.longitude + offset * 0.0012) if school else 0.0
            position_label = "Position de base (école)"
        delay_minutes = estimate_bus_delay(b, route, now=now) if route else None
        if delay_minutes is not None and delay_minutes > DELAY_THRESHOLD_MINUTES:
            delayed_buses += 1
        speed_kmh = b.speed_kmh or 0.0
        if speed_kmh > SPEED_LIMIT_KMH:
            speed_alerts += 1
        bus_markers.append(
            {
                "id": b.id,
                "code_bus": b.code_bus,
                "driver_name": b.driver_name,
                "capacity": b.capacity,
                "latitude": lat,
                "longitude": lon,
                "route_name": route.name if route else None,
                "position_label": position_label,
                "delay_minutes": delay_minutes,
                "speed_kmh": round(speed_kmh, 1),
                "is_speeding": speed_kmh > SPEED_LIMIT_KMH,
                "driver_connected": b.driver_connected,
            }
        )

    map_data = {
        "school": (
            {"name": school.name, "latitude": school.latitude, "longitude": school.longitude}
            if school else None
        ),
        "buses": bus_markers,
        "students": [
            {
                "matricule": s.matricule,
                "prenom": s.prenom,
                "postnom": s.postnom,
                "nom": s.nom,
                "address": s.address,
                "latitude": s.latitude,
                "longitude": s.longitude,
                "frozen": s.is_frozen,
            }
            for s in students
            if s.latitude is not None and s.longitude is not None
        ],
        "routes": [
            {
                "name": r.name,
                "bus": r.bus.code_bus if r.bus else "—",
                "driver": r.bus.driver_name if r.bus else "",
                "capacity": r.bus.capacity if r.bus else 0,
                "total_distance_km": r.total_distance_km,
                "estimated_duration_minutes": r.estimated_duration_minutes,
                "students_taken": r.students_taken,
                "students_remaining": r.students_remaining,
                "path_geometry": r.path_geometry or [],
                "path_source": "osrm" if r.path_geometry else "fallback",
                "stops": [
                    {
                        "order": rs.order,
                        "name": rs.stop.name,
                        "latitude": rs.stop.latitude,
                        "longitude": rs.stop.longitude,
                    }
                    for rs in r.stops.all().order_by("order")
                ],
            }
            for r in routes
        ],
    }

    incidents = Incident.objects.filter(bus__school=school, resolved=False).order_by("-timestamp")
    maintenance = BusMaintenance.objects.filter(bus__school=school)

    # Contrainte d'affectation : élèves avec position GPS mais sans bus attribué
    # (dont les cas de surcapacité) — affichés en bannière d'alerte prioritaire.
    unassigned_qs = _unassigned_students_with_gps(school)
    unassigned_count = unassigned_qs.count()
    overcapacity_count = unassigned_qs.filter(overcapacity_alert=True).count()

    # Compteurs globaux servis au chargement (mis à jour ensuite par /api/school/fleet/).
    total_taken = sum(r.students_taken for r in routes)
    total_remaining = sum(r.students_remaining for r in routes)
    stale_buses = sum(
        1
        for b in buses
        if b.driver_connected
        and b.last_position_at is not None
        and (timezone.now() - b.last_position_at).total_seconds() > 300
    )

    context = {
        "students": students,
        "buses": buses,
        "routes": routes,
        "school": school,
        "stats": _bus_statistics(school=school),
        "map_data": map_data,
        "student_form": StudentForm(),
        "bus_form": BusForm(),
        "incidents": incidents,
        "maintenance": maintenance,
        "maintenance_types": [t[0] for t in BusMaintenance.SERVICE_TYPES],
        "absent_today": absent_student_ids(),
        "unassigned_count": unassigned_count,
        "overcapacity_count": overcapacity_count,
        "live_taken": total_taken,
        "live_remaining": total_remaining,
        "live_incidents": incidents.count(),
        "live_unassigned": unassigned_count,
        "live_stale": stale_buses,
        "live_delayed": delayed_buses,
        "live_enroute": len(bus_markers),
        "live_speed_alerts": speed_alerts,
        "live_pickup": (
            round(total_taken / (total_taken + total_remaining) * 100, 1)
            if (total_taken + total_remaining) > 0
            else 0.0
        ),
    }
    return render(request, "routing/dashboard.html", context)


# ---------------------------------------------------------------- SMS d'urgence (école)


@login_required
def broadcast_sms_view(request):
    """Diffuse un SMS / notification d'urgence aux parents des élèves d'un bus donné."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    if request.method != "POST":
        return redirect("routing:dashboard")
    school = _user_school(request.user)
    bus = get_object_or_404(Bus, pk=request.POST.get("bus_id"), school=school)
    message = request.POST.get("message", "").strip()
    if not message:
        messages.error(request, "Le message SMS est vide.")
        return redirect("routing:dashboard")
    phones = set(
        Student.objects.filter(assigned_route__bus=bus)
        .exclude(parent_phone="")
        .values_list("parent_phone", flat=True)
    )
    sent = sum(1 for phone in phones if send_parent_sms(phone, message))
    messages.success(
        request, f"SMS d'urgence envoyé à {sent} parent(s) du bus {bus.code_bus}."
    )
    return redirect("routing:dashboard")


# ---------------------------------------------------------------- Maintenance flotte (école)


@login_required
def maintenance_add(request):
    """Enregistre une intervention d'entretien / révision sur un bus."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    if request.method == "POST":
        school = _user_school(request.user)
        bus = get_object_or_404(Bus, pk=request.POST.get("bus_id"), school=school)
        service_type = request.POST.get("service_type", "")
        if service_type not in {t[0] for t in BusMaintenance.SERVICE_TYPES}:
            messages.error(request, "Type d'entretien invalide.")
            return redirect("routing:dashboard")
        try:
            cost = Decimal(request.POST.get("cost") or 0)
        except (TypeError, ValueError, ArithmeticError):
            cost = Decimal(0)
        date_str = request.POST.get("date_effectuee") or timezone.localdate().isoformat()
        BusMaintenance.objects.create(
            bus=bus,
            service_type=service_type,
            cost=cost,
            date_effectuee=date_str,
            prochaine_echeance_km_ou_date=request.POST.get("prochaine_echeance", "").strip(),
        )
        messages.success(request, f"Entretien « {service_type} » enregistré pour {bus.code_bus}.")
    return redirect("routing:dashboard")


@login_required
def maintenance_delete(request, pk):
    """Supprime un enregistrement d'entretien."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    record = get_object_or_404(BusMaintenance, pk=pk)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Enregistrement d'entretien supprimé.")
    return redirect("routing:dashboard")


# ---------------------------------------------------------------- Replay GPS (école)


@login_required
def replay_view(request, bus_id):
    """Replay GPS : rejoue pas à pas le trajet parcouru par un bus sur une carte.
    Sélecteur de date + curseur de lecture (slider)."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    school = _user_school(request.user)
    bus = get_object_or_404(Bus, pk=bus_id, school=school)

    date_str = request.GET.get("date") or ""
    logs = bus.gps_logs.all()
    if date_str:
        try:
            target = datetime.strptime(date_str, "%Y-%m-%d").date()
            logs = bus.gps_logs.filter(timestamp__date=target)
        except ValueError:
            date_str = ""

    points = [
        {
            "lat": log.latitude,
            "lng": log.longitude,
            "speed": log.speed_kmh,
            "time": log.timestamp.strftime("%H:%M:%S"),
        }
        for log in logs
    ]
    available_dates = sorted(
        {log.timestamp.date().isoformat() for log in bus.gps_logs.all()}, reverse=True
    )
    context = {
        "bus": bus,
        "points": points,
        "points_json": json.dumps(points),
        "selected_date": date_str,
        "available_dates": available_dates,
    }
    return render(request, "routing/replay.html", context)


# ---------------------------------------------------------------- Élèves


@login_required
def student_add(request):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    form = StudentForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            student = form.save()
            if student.school is None:
                student.school = _user_school(request.user)
                student.save(update_fields=["school"])
            # Compte de connexion : identifiant = matricule, mot de passe = matricule.
            _ensure_student_account(student)
            # L'affectation est entièrement automatique (signal post_save →
            # recalculate_school_routes) : on récupère le résultat du moteur.
            student.refresh_from_db()
            route = student.assigned_route
            if route is not None and route.bus is not None:
                assigned_bus = route.bus
                messages.success(
                    request,
                    f"Élève « {student} » ajouté et affecté automatiquement au bus "
                    f"{assigned_bus.code_bus} (chauffeur : {assigned_bus.driver_name or '—'}).",
                )
            else:
                messages.success(
                    request,
                    f"Élève « {student} » ajouté. Compte de connexion : {student.matricule} / "
                    f"{student.matricule} (mot de passe à changer à la première connexion, "
                    "puis position du domicile à définir).",
                )
            return redirect("routing:dashboard")
        messages.error(request, "Le formulaire contient des erreurs.")
    return render(request, "routing/student_form.html", {"form": form, "title": "Ajouter un élève"})


@login_required
def student_edit(request, pk):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    student = get_object_or_404(Student, pk=pk)
    form = StudentForm(request.POST or None, instance=student)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, f"Élève « {form.instance} » mis à jour.")
            return redirect("routing:dashboard")
        messages.error(request, "Le formulaire contient des erreurs.")
    return render(request, "routing/student_form.html", {"form": form, "title": f"Modifier {student}"})


@login_required
def student_delete(request, pk):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        matricule = student.matricule
        student.delete()
        messages.success(request, f"Élève {matricule} supprimé.")
    return redirect("routing:dashboard")


@login_required
def toggle_freeze_student(request, pk):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        student.is_frozen = not student.is_frozen
        student.save(update_fields=["is_frozen"])
        state = "gelé" if student.is_frozen else "dégelé"
        messages.success(request, f"Élève {student.matricule} {state}.")
    return redirect("routing:dashboard")


@login_required
def student_reset_password(request, pk):
    """Réinitialise le mot de passe du compte élève/parent
    (réservé à l'école de l'élève)."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    student = get_object_or_404(Student, pk=pk)
    if student.school_id != _user_school(request.user).pk:
        messages.error(request, "Cet élève ne fait pas partie de votre école.")
        return redirect("routing:dashboard")
    if request.method == "POST":
        p1 = request.POST.get("password1", "")
        p2 = request.POST.get("password2", "")
        if len(p1) < 4:
            messages.error(request, "Le mot de passe doit contenir au moins 4 caractères.")
        elif p1 != p2:
            messages.error(request, "Les deux mots de passe ne correspondent pas.")
        else:
            user = _ensure_student_account(student)
            user.set_password(p1)
            user.save(update_fields=["password"])
            # Sécurité : le parent devra personnaliser ce mot de passe
            # à sa prochaine connexion (onboarding déjà en place).
            student.must_change_password = True
            student.save(update_fields=["must_change_password"])
            messages.success(
                request,
                f"Mot de passe de l'élève {student.matricule} réinitialisé. "
                "Il devra le personnaliser à sa prochaine connexion.",
            )
    return redirect("routing:dashboard")


# ---------------------------------------------------------------- Bus


@login_required
def bus_add(request):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    form = BusForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            bus = form.save()
            # Rattachement à l'école de l'utilisateur (sinon le bus est invisible
            # dans le tableau de bord qui filtre par école).
            if bus.school is None:
                bus.school = _user_school(request.user)
                bus.save(update_fields=["school"])
            # Compte de connexion du chauffeur : identifiant = code du bus,
            # mot de passe = code du bus (à changer à la première connexion).
            _ensure_driver_account(bus)
            messages.success(
                request,
                f"Bus « {bus.code_bus} » ajouté. Compte chauffeur : {bus.code_bus} / "
                f"{bus.code_bus} (mot de passe à changer à la première connexion).",
            )
            return redirect("routing:dashboard")
        messages.error(request, "Le formulaire contient des erreurs.")
    return render(request, "routing/bus_form.html", {"form": form, "title": "Ajouter un bus"})


@login_required
def bus_edit(request, pk):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    bus = get_object_or_404(Bus, pk=pk)
    form = BusForm(request.POST or None, instance=bus)
    if request.method == "POST":
        if form.is_valid():
            bus = form.save()
            # Rattachement à l'école si le bus était orphelin (créé avant la correction).
            if bus.school is None:
                bus.school = _user_school(request.user)
                bus.save(update_fields=["school"])
            # Récupération : si le bus n'a pas encore de compte chauffeur, on le crée.
            if bus.driver_user_id is None:
                _ensure_driver_account(bus)
            messages.success(request, f"Bus « {bus.code_bus} » mis à jour.")
            return redirect("routing:dashboard")
        messages.error(request, "Le formulaire contient des erreurs.")
    return render(request, "routing/bus_form.html", {"form": form, "title": f"Modifier {bus}"})


@login_required
def bus_delete(request, pk):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    bus = get_object_or_404(Bus, pk=pk)
    if request.method == "POST":
        code = bus.code_bus
        bus.delete()
        messages.success(request, f"Bus {code} supprimé.")
    return redirect("routing:dashboard")


@login_required
def bus_reset_password(request, pk):
    """Réinitialise le mot de passe du compte chauffeur d'un bus
    (réservé à l'école du bus)."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    bus = get_object_or_404(Bus, pk=pk)
    if bus.school_id != _user_school(request.user).pk:
        messages.error(request, "Ce bus ne fait pas partie de votre école.")
        return redirect("routing:dashboard")
    if request.method == "POST":
        p1 = request.POST.get("password1", "")
        p2 = request.POST.get("password2", "")
        if len(p1) < 4:
            messages.error(request, "Le mot de passe doit contenir au moins 4 caractères.")
        elif p1 != p2:
            messages.error(request, "Les deux mots de passe ne correspondent pas.")
        else:
            user = _ensure_driver_account(bus)
            user.set_password(p1)
            user.save(update_fields=["password"])
            # Sécurité : le chauffeur devra personnaliser ce mot de passe
            # à sa prochaine connexion.
            bus.must_change_password = True
            bus.save(update_fields=["must_change_password"])
            messages.success(
                request,
                f"Mot de passe du chauffeur du bus {bus.code_bus} réinitialisé. "
                "Il devra le personnaliser à sa prochaine connexion.",
            )
    return redirect("routing:dashboard")


# ---------------------------------------------------------------- Statistiques


@login_required
def statistics_view(request):
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    school = _user_school(request.user)
    context = {"stats": _bus_statistics(school=school)}
    return render(request, "routing/statistics.html", context)


# ---------------------------------------------------------------- Administration générale (écoles)


def _require_superuser(request):
    """Redirige vers le dashboard si l'utilisateur n'est pas superutilisateur."""
    if not request.user.is_superuser:
        messages.error(request, "Accès réservé à l'administration générale.")
        return redirect("routing:dashboard")
    return None


@login_required
def schools_admin_view(request):
    """Administration générale : liste de toutes les écoles avec leurs statistiques."""
    blocked = _require_superuser(request)
    if blocked:
        return blocked

    schools = School.objects.prefetch_related("students", "buses", "routes").order_by("code_ecole")
    rows = []
    total_taken = total_remaining = 0
    for school in schools:
        routes = school.routes.all()
        taken = sum(r.students_taken for r in routes)
        remaining = sum(r.students_remaining for r in routes)
        assigned = taken + remaining
        pickup_rate = round(taken / assigned * 100, 1) if assigned else 0.0
        total_taken += taken
        total_remaining += remaining
        rows.append(
            {
                "school": school,
                "students_count": school.students.count(),
                "buses_count": school.buses.count(),
                "routes_count": school.routes.count(),
                "taken": taken,
                "remaining": remaining,
                "pickup_rate": pickup_rate,
                "last_route": school.routes.order_by("-created_at").first(),
            }
        )
    active_count = schools.filter(is_active=True).count()
    schools_count = schools.count()
    total_students = sum(r["students_count"] for r in rows)
    total_buses = sum(r["buses_count"] for r in rows)
    total_routes = sum(r["routes_count"] for r in rows)
    total_assigned = total_taken + total_remaining
    overall_pickup = round(total_taken / total_assigned * 100, 1) if total_assigned else 0.0

    chart_data = {
        "codes": [r["school"].code_ecole for r in rows],
        "students": [r["students_count"] for r in rows],
        "buses": [r["buses_count"] for r in rows],
        "routes": [r["routes_count"] for r in rows],
        "pickup": [r["pickup_rate"] for r in rows],
        "active": active_count,
        "inactive": schools_count - active_count,
    }
    context = {
        "rows": rows,
        "schools_count": schools_count,
        "active_count": active_count,
        "inactive_count": schools_count - active_count,
        "total_students": total_students,
        "total_buses": total_buses,
        "total_routes": total_routes,
        "total_taken": total_taken,
        "total_remaining": total_remaining,
        "overall_pickup": overall_pickup,
        "chart_data": json.dumps(chart_data),
    }
    return render(request, "routing/schools_admin.html", context)


@login_required
def toggle_school_active(request, pk):
    """Active / désactive une école (réservé à l'administration générale)."""
    blocked = _require_superuser(request)
    if blocked:
        return blocked
    school = get_object_or_404(School, pk=pk)
    if request.method == "POST":
        school.is_active = not school.is_active
        school.save(update_fields=["is_active"])
        state = "activée" if school.is_active else "désactivée"
        messages.success(request, f"École {school.code_ecole} {state}.")
    return redirect("routing:schools_admin")


@login_required
def school_delete(request, pk):
    """Supprime une école et toutes ses données (réservé à l'administration générale)."""
    blocked = _require_superuser(request)
    if blocked:
        return blocked
    school = get_object_or_404(School, pk=pk)
    if request.method == "POST":
        code = school.code_ecole
        admin_user = school.user
        school.delete()  # CASCADE : élèves, bus, routes, trajets
        if admin_user is not None:
            admin_user.delete()  # compte admin dédié à l'école
        messages.success(
            request,
            f"École {code} supprimée avec toutes ses données "
            "(élèves, bus, trajets, compte administrateur).",
        )
    return redirect("routing:schools_admin")


@login_required
def school_reset_password(request, pk):
    """Réinitialise le mot de passe du compte administrateur d'une école
    (réservé à l'administration générale)."""
    blocked = _require_superuser(request)
    if blocked:
        return blocked
    school = get_object_or_404(School, pk=pk)
    if request.method == "POST":
        p1 = request.POST.get("password1", "")
        p2 = request.POST.get("password2", "")
        if len(p1) < 4:
            messages.error(request, "Le mot de passe doit contenir au moins 4 caractères.")
        elif p1 != p2:
            messages.error(request, "Les deux mots de passe ne correspondent pas.")
        else:
            user = school.user
            if user is None:
                # L'école n'a pas encore de compte admin (cas rare) : on le crée.
                user = User.objects.create_user(username=school.code_ecole, password=p1)
                school.user = user
                school.save(update_fields=["user"])
            else:
                user.set_password(p1)
                user.save(update_fields=["password"])
            messages.success(
                request,
                f"Mot de passe de l'école {school.code_ecole} réinitialisé. "
                "Communiquez le nouveau mot de passe à l'école.",
            )
    return redirect("routing:schools_admin")


# ---------------------------------------------------------------- Parent / Élève


@login_required
def parent_dashboard_view(request):
    """Interface parent/élève : suivi du bus en temps réel et statut du ramassage."""
    if _pending_password_change(request.user):
        return redirect("routing:change_password")
    student = _student_for_user(request.user)
    if student is None:
        messages.error(request, "Aucun élève n'est associé à votre compte.")
        return redirect("routing:dashboard")
    if student.latitude is None or student.longitude is None:
        # Onboarding : la maison doit d'abord être positionnée sur la carte.
        return redirect("routing:parent_setup_home")

    route = student.assigned_route if student.assigned_route else None
    bus = route.bus if route else None

    absent_today = Absence.objects.filter(
        student=student, date=timezone.localdate()
    ).exists()
    context = {
        "student": student,
        "bus": bus,
        "absent_today": absent_today,
        "student_data": {
            "name": " ".join(filter(None, [student.prenom, student.postnom, student.nom])),
            "latitude": student.latitude,
            "longitude": student.longitude,
        },
    }
    return render(request, "routing/parent.html", context)


@login_required
def parent_setup_home_view(request):
    """Étape 2 de l'onboarding : le parent positionne le domicile de l'élève
    sur une carte Leaflet (clic sur la carte ou bouton « position GPS actuelle »).
    L'accès à /parent/ reste bloqué tant que la maison n'est pas positionnée."""
    student = _student_for_user(request.user)
    if student is None:
        messages.error(request, "Aucun élève n'est associé à votre compte.")
        return redirect("routing:dashboard")

    # Déjà positionné → direction l'interface parent.
    if student.latitude is not None and student.longitude is not None:
        return redirect("routing:parent")

    if request.method == "POST":
        try:
            latitude = float(request.POST.get("latitude"))
            longitude = float(request.POST.get("longitude"))
        except (TypeError, ValueError):
            latitude = longitude = None
        if (
            latitude is not None
            and longitude is not None
            and -90 <= latitude <= 90
            and -180 <= longitude <= 180
        ):
            student.latitude = latitude
            student.longitude = longitude
            student.save(update_fields=["latitude", "longitude"])
            # Le signal post_save a relancé le moteur central d'affectation :
            # l'élève est immédiatement rattaché au bus le plus proche (capacité
            # libre) et le parent voit le bus et le chauffeur attribués.
            student.refresh_from_db()
            route = student.assigned_route
            if route is not None and route.bus is not None:
                assigned_bus = route.bus
                messages.success(
                    request,
                    f"Domicile enregistré avec succès. 🚌 Bus attribué automatiquement : "
                    f"{assigned_bus.code_bus} (chauffeur : {assigned_bus.driver_name or '—'}). "
                    f"Bienvenue {student.prenom} ! 🎉",
                )
            elif student.overcapacity_alert:
                messages.warning(
                    request,
                    "Domicile enregistré, mais tous les bus proches sont complets. "
                    "L'administration de l'école a été alertée (surcapacité).",
                )
            else:
                messages.success(
                    request, f"Domicile enregistré avec succès. Bienvenue {student.prenom} ! 🎉"
                )
            return redirect("routing:parent")
        messages.error(
            request,
            "Coordonnées invalides — placez le marqueur sur la carte ou utilisez votre position GPS.",
        )

    # Centre de la carte : l'école de l'élève si connue, sinon Lubumbashi (région démo).
    if student.school is not None:
        center = {
            "latitude": student.school.latitude,
            "longitude": student.school.longitude,
        }
    else:
        center = {"latitude": -11.6647, "longitude": 27.4794}
    context = {"student": student, "map_center": center}
    return render(request, "routing/setup_home.html", context)


@login_required
def parent_toggle_absence(request):
    """Le parent signale / annule l'absence de l'élève pour aujourd'hui.
    Un élève absent est exclu des tournées du jour (aucun arrêt planifié)."""
    student = _student_for_user(request.user)
    if student is None:
        messages.error(request, "Aucun élève n'est associé à votre compte.")
        return redirect("routing:parent")
    if request.method == "POST":
        today = timezone.localdate()
        absence = Absence.objects.filter(student=student, date=today).first()
        if absence:
            absence.delete()
            messages.success(request, "Absence annulée — votre enfant sera pris en charge.")
        else:
            reason = request.POST.get("reason", "").strip()
            Absence.objects.get_or_create(student=student, date=today, defaults={"reason": reason})
            messages.success(
                request, "Absence signalée pour aujourd'hui — l'arrêt du bus sera ignoré."
            )
    return redirect("routing:parent")


@login_required
def parent_live(request):
    """API JSON (/api/parent/live/) : position GPS du bus, tracé complet,
    domicile de l'élève, statut du trajet et vitesse."""
    student = _student_for_user(request.user)
    if student is None:
        return JsonResponse({"error": "Aucun élève associé."}, status=404)

    route = student.assigned_route if student.assigned_route else None
    bus = route.bus if route else None
    has_gps = bus is not None and bus.last_latitude is not None and bus.last_longitude is not None
    # Signal GPS « en direct » : reçu il y a moins de 60 secondes.
    live = bool(
        has_gps
        and bus.last_position_at is not None
        and (timezone.now() - bus.last_position_at).total_seconds() <= 60
    )
    if route is None:
        status = "no_route"
    elif live:
        status = "en_route"
    else:
        status = "at_depot"
    speed_kmh = bus.speed_kmh if bus else 0.0

    return JsonResponse(
        {
            "bus_lat": bus.last_latitude if has_gps else None,
            "bus_lng": bus.last_longitude if has_gps else None,
            "has_gps": has_gps,
            "bus_code": bus.code_bus if bus else None,
            "path_geometry": route.path_geometry if route and route.path_geometry else [],
            "route_name": route.name if route else None,
            "status": status,
            "speed_kmh": round(speed_kmh, 1) if has_gps else None,
            "is_speeding": speed_kmh > SPEED_LIMIT_KMH,
            "eta_minutes": estimate_remaining_minutes(bus, route) if route else None,
            "student_lat": student.latitude,
            "student_lng": student.longitude,
            "is_picked_up": student.is_taken,
            "absent_today": Absence.objects.filter(
                student=student, date=timezone.localdate()
            ).exists(),
            "school_lat": route.school.latitude if route and route.school else None,
            "school_lng": route.school.longitude if route and route.school else None,
            "position_at": (
                bus.last_position_at.isoformat() if bus and bus.last_position_at else None
            ),
        }
    )


# ---------------------------------------------------------------- API localisation temps réel


def _can_view_bus(user, bus):
    """Accès à la localisation d'un bus : superuser, école du bus, chauffeur
    du bus, ou parent d'un élève affecté à ce bus."""
    if user.is_superuser:
        return True
    driven = getattr(user, "driven_bus", None)
    if driven is not None and driven.pk == bus.pk:
        return True
    school = getattr(user, "school", None)
    if school is not None and school.pk == bus.school_id:
        return True
    student = _student_for_user(user)
    if (
        student is not None
        and student.assigned_route_id is not None
        and student.assigned_route.bus_id == bus.pk
    ):
        return True
    return False


@login_required
def bus_location_api(request, bus_id):
    """API JSON (/api/bus/<bus_id>/location/) : position GPS temps réel du bus,
    feuille de route active et statut du trajet.

    Retourne :
      - latitude / longitude / speed_kmh / last_updated : position actuelle ;
      - route : nom, tracé routier (path_geometry) et arrêts ordonnés ;
      - status : 'no_route' (aucune feuille de route), 'at_depot' (bus à
        l'école, en attente du départ) ou 'en_route' (GPS en direct).

    Accès : école du bus, chauffeur du bus ou parent d'un élève affecté.
    """
    bus = get_object_or_404(Bus, pk=bus_id)
    if not _can_view_bus(request.user, bus):
        return JsonResponse({"error": "Accès refusé."}, status=403)

    route = bus.routes.order_by("-created_at").first()
    live = bool(
        bus.last_latitude is not None
        and bus.last_longitude is not None
        and bus.last_position_at is not None
        and (timezone.now() - bus.last_position_at).total_seconds() <= 60
    )
    if route is None:
        status = "no_route"
    elif live:
        status = "en_route"
    else:
        status = "at_depot"

    speed_kmh = bus.speed_kmh or 0.0

    return JsonResponse(
        {
            "bus_id": bus.pk,
            "code_bus": bus.code_bus,
            "driver_name": bus.driver_name or "",
            "latitude": bus.last_latitude if live else None,
            "longitude": bus.last_longitude if live else None,
            "speed_kmh": round(speed_kmh, 1) if live else None,
            "is_speeding": speed_kmh > SPEED_LIMIT_KMH,
            "last_updated": (
                bus.last_position_at.isoformat() if bus.last_position_at else None
            ),
            "status": status,
            "is_active": route is not None,
            "delay_minutes": estimate_bus_delay(bus, route) if route else None,
            "eta_minutes": estimate_remaining_minutes(bus, route) if route else None,
            "route": (
                {
                    "name": route.name,
                    "path_geometry": route.path_geometry or [],
                    "stops": [
                        {
                            "order": rs.order,
                            "name": rs.stop.name,
                            "latitude": rs.stop.latitude,
                            "longitude": rs.stop.longitude,
                        }
                        for rs in route.stops.order_by("order")
                    ],
                    "students_taken": route.students_taken,
                    "students_remaining": route.students_remaining,
                }
                if route
                else None
            ),
        }
    )


@login_required
def school_fleet_api(request):
    """API JSON (/api/school/fleet/) : état temps réel de la flotte pour le
    tableau de bord de l'école (positions des bus + compteurs globaux).

    Seuls les bus en service dont le chauffeur est connecté apparaissent
    (position récente < 5 min) — cohérent avec la carte initiale : un bus
    apparaît/disparaît automatiquement à la connexion/déconnexion.
    """
    blocked = _school_scope_guard(request)
    if blocked:
        return JsonResponse({"error": "Accès refusé."}, status=403)
    school = _user_school(request.user)
    if school is None:
        return JsonResponse({"buses": [], "counters": {}})

    now = timezone.now()
    buses = []
    stale_buses = 0
    delayed_buses = 0
    for offset, b in enumerate(
        Bus.objects.filter(school=school, is_in_service=True).order_by("code_bus")
    ):
        stale = (
            b.last_position_at is not None
            and (now - b.last_position_at).total_seconds() > 300
        )
        if b.driver_connected and stale:
            stale_buses += 1  # chauffeur connecté mais signal GPS perdu (retard possible)
        if not b.driver_connected or stale:
            continue
        route = b.routes.order_by("-created_at").first()
        has_gps = b.last_latitude is not None and b.last_longitude is not None
        if has_gps:
            lat, lon = b.last_latitude, b.last_longitude
        else:
            # Chauffeur connecté mais position pas encore reçue : base = école,
            # cohérent avec le rendu initial de la carte du tableau de bord.
            lat = school.latitude + offset * 0.0012
            lon = school.longitude + offset * 0.0012
        delay_minutes = estimate_bus_delay(b, route, now=now) if route else None
        if delay_minutes is not None and delay_minutes > DELAY_THRESHOLD_MINUTES:
            delayed_buses += 1
        speed_kmh = b.speed_kmh or 0.0
        buses.append(
            {
                "id": b.pk,
                "code_bus": b.code_bus,
                "driver_name": b.driver_name or "",
                "capacity": b.capacity,
                "latitude": lat,
                "longitude": lon,
                "at_base": not has_gps,
                "route_name": route.name if route else None,
                "taken": route.students_taken if route else 0,
                "remaining": route.students_remaining if route else 0,
                "delay_minutes": delay_minutes,
                "speed_kmh": round(speed_kmh, 1),
                "is_speeding": speed_kmh > SPEED_LIMIT_KMH,
            }
        )

    routes = Route.objects.filter(school=school)
    taken = sum(r.students_taken for r in routes)
    remaining = sum(r.students_remaining for r in routes)
    assigned = taken + remaining
    unassigned_qs = _unassigned_students_with_gps(school)
    return JsonResponse(
        {
            "buses": buses,
            "counters": {
                "taken": taken,
                "remaining": remaining,
                "assigned": assigned,
                "pickup_rate": round(taken / assigned * 100, 1) if assigned else 0.0,
                "incidents": Incident.objects.filter(
                    bus__school=school, resolved=False
                ).count(),
                "unassigned": unassigned_qs.count(),
                "overcapacity": unassigned_qs.filter(overcapacity_alert=True).count(),
                "stale_buses": stale_buses,
                "delayed_buses": delayed_buses,
                "buses_en_route": len(buses),
                "speed_alerts": sum(
                    1 for b in buses if b["is_speeding"]
                ),
            },
        }
    )


# ---------------------------------------------------------------- Chauffeur


def _route_students_in_order(route):
    """Retourne les élèves du trajet, ordonnés selon l'ordre des arrêts.

    Repli : si le trajet n'a pas d'arrêts (ou si des élèves ne sont couverts
    par aucun arrêt), ils sont ajoutés à la fin dans un ordre stable.
    """
    route_students = list(route.students.all())
    if not route_students:
        return []
    ordered = []
    for rs in route.stops.order_by("order"):
        stop = rs.stop
        best = min(
            route_students,
            key=lambda s: haversine_km(
                s.latitude, s.longitude, stop.latitude, stop.longitude
            ),
        )
        if best not in ordered:
            ordered.append(best)
    for s in route_students:
        if s not in ordered:
            ordered.append(s)
    return ordered


@login_required
def driver_view(request):
    """Interface chauffeur : GPS en direct, trajet ciblé et ramassage."""
    if _pending_password_change(request.user):
        return redirect("routing:change_password")
    bus = getattr(request.user, "driven_bus", None)
    if bus is None:
        messages.error(request, "Aucun bus n'est assigné à votre compte chauffeur.")
        return redirect("routing:dashboard")

    route = bus.routes.order_by("-created_at").first()
    absent = absent_student_ids()
    # Tous les élèves du trajet (pour la carte) ; les absents du jour y figurent
    # avec le drapeau « absent » (marqueur rouge) mais pas sur la feuille de route.
    route_students = _route_students_in_order(route) if route else []
    students = [s for s in route_students if s.id not in absent]

    taken = sum(1 for s in students if s.is_taken)
    total = len(students)
    remaining = total - taken

    start = None
    if bus.last_latitude is not None and bus.last_longitude is not None:
        start = [bus.last_latitude, bus.last_longitude]
    elif route and route.path_geometry:
        start = route.path_geometry[0]
    elif route and route.school:
        start = [route.school.latitude, route.school.longitude]

    map_data = {
        "bus": {
            "code_bus": bus.code_bus,
            "latitude": start[0] if start else 0.0,
            "longitude": start[1] if start else 0.0,
        },
        "school": (
            {
                "name": bus.school.name,
                "latitude": bus.school.latitude,
                "longitude": bus.school.longitude,
            }
            if bus.school else None
        ),
        "route": (
            {
                "name": route.name,
                "total_distance_km": route.total_distance_km,
                "estimated_duration_minutes": route.estimated_duration_minutes,
                "path_geometry": route.path_geometry or [],
            }
            if route
            else None
        ),
        "students": [
            {
                "id": s.id,
                "matricule": s.matricule,
                "name": " ".join(filter(None, [s.prenom, s.postnom, s.nom])),
                "latitude": s.latitude,
                "longitude": s.longitude,
                "taken": s.is_taken,
                "absent": s.id in absent,
                "order": idx,
            }
            for idx, s in enumerate(route_students)
        ],
    }

    context = {
        "bus": bus,
        "route": route,
        "students": students,
        "taken": taken,
        "remaining": remaining,
        "total": total,
        "map_data": map_data,
    }
    return render(request, "routing/driver.html", context)


@login_required
def driver_update_position(request):
    """API AJAX : enregistre la position GPS temps réel du bus du chauffeur."""
    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée."}, status=405)

    bus = getattr(request.user, "driven_bus", None)
    if bus is None:
        return JsonResponse({"error": "Aucun bus assigné à ce compte."}, status=403)

    try:
        latitude = float(request.POST.get("latitude"))
        longitude = float(request.POST.get("longitude"))
    except (TypeError, ValueError):
        return JsonResponse({"error": "Coordonnées invalides."}, status=400)
    if not (-90.0 <= latitude <= 90.0) or not (-180.0 <= longitude <= 180.0):
        return JsonResponse({"error": "Coordonnées hors limites."}, status=400)

    # --- Vitesse : mesurée par l'appareil du chauffeur, sinon estimée à partir
    # du point GPS précédent (Haversine / temps écoulé).
    try:
        speed_kmh = float(request.POST.get("speed") or 0.0)
    except (TypeError, ValueError):
        speed_kmh = 0.0
    if speed_kmh <= 0:
        prev = bus.gps_logs.order_by("-timestamp").first()
        if prev is not None:
            delta_s = (timezone.now() - prev.timestamp).total_seconds()
            if delta_s > 0:
                dist_km = haversine_km(prev.latitude, prev.longitude, latitude, longitude)
                speed_kmh = round(dist_km / (delta_s / 3600.0), 1)
    speed_kmh = max(0.0, speed_kmh)

    bus.last_latitude = latitude
    bus.last_longitude = longitude
    bus.last_position_at = timezone.now()
    # Battement de cœur : tant que le chauffeur envoie sa position,
    # le bus reste visible en direct sur la carte de l'école.
    bus.driver_connected = True
    bus.speed_kmh = speed_kmh
    bus.save(
        update_fields=[
            "last_latitude",
            "last_longitude",
            "last_position_at",
            "driver_connected",
            "speed_kmh",
        ]
    )

    # --- Historique GPS (replay) : enregistre chaque point de passage.
    GPSLog.objects.create(bus=bus, latitude=latitude, longitude=longitude, speed_kmh=speed_kmh)

    # --- Sécurité routière : alerte en cas d'excès de vitesse.
    speed_limit = SPEED_LIMIT_KMH
    if speed_kmh > speed_limit:
        has_recent = Incident.objects.filter(
            bus=bus,
            type_incident=Incident.TYPE_VITESSE,
            resolved=False,
            timestamp__gte=timezone.now() - timedelta(minutes=10),
        ).exists()
        if not has_recent:
            Incident.objects.create(
                bus=bus,
                type_incident=Incident.TYPE_VITESSE,
                description=(
                    f"Vitesse {speed_kmh:.0f} km/h dépassant la limite de {speed_limit:.0f} km/h."
                ),
            )

    # Alerte SMS aux parents : si le bus franchit le seuil de 100 m d'un élève
    # non encore pris, on prévient le parent une seule fois (parent_notified).
    approaching = Student.objects.filter(
        assigned_route__bus=bus, is_taken=False, parent_phone__gt=""
    )
    for student in approaching:
        distance_m = (
            haversine_km(latitude, longitude, student.latitude, student.longitude) * 1000
        )
        if distance_m <= 100 and not student.parent_notified:
            student.parent_notified = True
            student.save(update_fields=["parent_notified"])
            send_parent_sms(
                student.parent_phone,
                f"ATTENTION : le bus {bus.code_bus} est à moins de 100 m de chez vous. "
                f"Préparez {student.prenom} !",
            )

    return JsonResponse(
        {"ok": True, "latitude": latitude, "longitude": longitude, "speed_kmh": speed_kmh}
    )


@login_required
def driver_pickup_student(request, pk):
    """API AJAX : marque un élève comme pris en charge (ramassage confirmé)."""
    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée."}, status=405)

    bus = getattr(request.user, "driven_bus", None)
    student = get_object_or_404(Student, pk=pk)
    if bus is None or student.assigned_route is None or student.assigned_route.bus_id != bus.id:
        return JsonResponse({"error": "Élève non autorisé pour ce bus."}, status=403)
    if student.is_taken:
        return JsonResponse({"error": "Élève déjà pris en charge."}, status=400)

    student.is_taken = True
    student.save(update_fields=["is_taken"])

    route = student.assigned_route
    route.students_taken += 1
    route.students_remaining = max(0, route.students_remaining - 1)
    route.save(update_fields=["students_taken", "students_remaining"])

    return JsonResponse(
        {
            "ok": True,
            "student": student.matricule,
            "taken": route.students_taken,
            "remaining": route.students_remaining,
            "total": route.students.count(),
        }
    )


@login_required
def driver_report_incident(request):
    """API AJAX : le chauffeur signale un incident (SOS préremplis).
    L'incident apparaît immédiatement sur le tableau de bord de l'école."""
    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée."}, status=405)
    bus = getattr(request.user, "driven_bus", None)
    if bus is None:
        return JsonResponse({"error": "Aucun bus assigné à ce compte."}, status=403)
    type_incident = request.POST.get("type_incident", "").strip()
    valid_types = {t[0] for t in Incident.INCIDENT_TYPES}
    if type_incident not in valid_types:
        return JsonResponse({"error": "Type d'incident invalide."}, status=400)
    incident = Incident.objects.create(
        bus=bus,
        type_incident=type_incident,
        description=request.POST.get("description", "").strip(),
    )
    return JsonResponse(
        {
            "ok": True,
            "id": incident.id,
            "type_incident": incident.type_incident,
            "bus": bus.code_bus,
        }
    )


@login_required
def driver_route_api(request):
    """API JSON : feuille de route du chauffeur (bus, tracé, élèves à récupérer).
    Les élèves déclarés absents aujourd'hui sont exclus."""
    bus = getattr(request.user, "driven_bus", None)
    if bus is None:
        return JsonResponse({"error": "Aucun bus assigné."}, status=403)
    route = bus.routes.order_by("-created_at").first()
    absent = absent_student_ids()
    route_students = _route_students_in_order(route) if route else []
    students = [s for s in route_students if s.id not in absent]
    taken = sum(1 for s in students if s.is_taken)
    return JsonResponse(
        {
            "bus": {"code_bus": bus.code_bus},
            "route": (
                {"name": route.name, "path_geometry": route.path_geometry or []}
                if route
                else None
            ),
            "students": [
                {
                    "id": s.id,
                    "matricule": s.matricule,
                    "name": " ".join(filter(None, [s.prenom, s.postnom, s.nom])),
                    "latitude": s.latitude,
                    "longitude": s.longitude,
                    "taken": s.is_taken,
                    "absent": s.id in absent,
                }
                for s in route_students
            ],
            "taken": taken,
            "remaining": len(students) - taken,
            "total": len(students),
        }
    )


def _parse_iso_timestamp(value):
    """Analyse un horodatage reçu du téléphone : ISO 8601 (avec ou sans fuseau)
    ou timestamp epoch (secondes ou millisecondes). Retourne un datetime aware,
    ou None si la valeur est absente ou illisible."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        # Millisecondes (> 1e11) ou secondes.
        seconds = value / 1000.0 if value > 1e11 else value
        return datetime.fromtimestamp(seconds, tz=timezone.utc)
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


@login_required
def sync_offline_data(request):
    """API JSON (/api/sync-offline-data/) : resynchronisation groupée des données
    accumulées par le téléphone du chauffeur pendant une coupure réseau.

    Reçoit un tableau de points GPS et/ou d'incidents, les insère en masse
    (bulk_create) et met à jour le dernier état connu du bus sans perte de
    données : une position plus ancienne que celle déjà en base n'écrase jamais
    l'état le plus récent.

    Corps attendu : {"points": [{latitude, longitude, speed, timestamp}],
                     "incidents": [{type_incident, description, timestamp}]}
    """
    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée."}, status=405)

    bus = getattr(request.user, "driven_bus", None)
    if bus is None:
        return JsonResponse({"error": "Aucun bus assigné à ce compte."}, status=403)

    try:
        payload = json.loads(request.body or b"{}")
    except (TypeError, ValueError):
        return JsonResponse({"error": "Corps JSON invalide."}, status=400)
    points = payload.get("points") or []
    incidents = payload.get("incidents") or []
    if not isinstance(points, list) or not isinstance(incidents, list):
        return JsonResponse(
            {"error": "points et incidents doivent être des tableaux."}, status=400
        )

    valid_types = {t[0] for t in Incident.INCIDENT_TYPES}
    now = timezone.now()
    max_batch = 5000  # garde-fou : limite par requête

    gps_objects = []
    incident_objects = []
    last_point = None  # (timestamp, latitude, longitude, vitesse) le plus récent du lot

    for item in points[:max_batch]:
        try:
            latitude = float(item.get("latitude"))
            longitude = float(item.get("longitude"))
            speed = float(item.get("speed") or 0.0)
        except (TypeError, ValueError, AttributeError):
            continue  # point illisible : on l'ignore, pas de blocage du lot
        if not (-90.0 <= latitude <= 90.0) or not (-180.0 <= longitude <= 180.0):
            continue
        timestamp = _parse_iso_timestamp(item.get("timestamp")) or now
        speed = max(0.0, speed)
        gps_objects.append(
            GPSLog(
                bus=bus,
                latitude=latitude,
                longitude=longitude,
                speed_kmh=speed,
                timestamp=timestamp,
            )
        )
        if last_point is None or timestamp > last_point[0]:
            last_point = (timestamp, latitude, longitude, speed)

    for item in incidents[:max_batch]:
        type_incident = str(item.get("type_incident") or "").strip()
        if type_incident not in valid_types:
            continue
        incident_objects.append(
            Incident(
                bus=bus,
                type_incident=type_incident,
                description=str(item.get("description") or "").strip(),
                timestamp=_parse_iso_timestamp(item.get("timestamp")) or now,
            )
        )

    if not gps_objects and not incident_objects:
        return JsonResponse({"ok": True, "synced_points": 0, "synced_incidents": 0})

    with transaction.atomic():
        GPSLog.objects.bulk_create(gps_objects)
        Incident.objects.bulk_create(incident_objects)
        # Dernier état connu du bus : on n'écrase jamais une position plus récente.
        if last_point is not None and (
            bus.last_position_at is None or last_point[0] > bus.last_position_at
        ):
            bus.last_latitude = last_point[1]
            bus.last_longitude = last_point[2]
            bus.last_position_at = last_point[0]
            bus.speed_kmh = last_point[3]
        # Battement de cœur : le bus reste visible en direct sur la carte de l'école.
        bus.driver_connected = True
        bus.save(
            update_fields=[
                "last_latitude",
                "last_longitude",
                "last_position_at",
                "speed_kmh",
                "driver_connected",
            ]
        )

    return JsonResponse(
        {
            "ok": True,
            "synced_points": len(gps_objects),
            "synced_incidents": len(incident_objects),
        }
    )


# ---------------------------------------------------------------- Imports Excel


def _read_uploaded_excel(request, required_columns):
    """Lit le fichier Excel uploadé et renvoie (index des colonnes, lignes)."""
    file = request.FILES.get("file")
    if file is None:
        messages.error(request, "Aucun fichier sélectionné.")
        return None
    if not file.name.lower().endswith(".xlsx"):
        messages.error(request, "Le fichier doit être au format .xlsx.")
        return None
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Lecture du fichier Excel impossible : %s", exc)
        messages.error(request, "Impossible de lire le fichier Excel.")
        return None

    rows = workbook.active.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        messages.error(request, "Le fichier Excel est vide.")
        return None

    columns = {
        str(cell).strip().lower(): i
        for i, cell in enumerate(header)
        if cell is not None
    }
    missing = [col for col in required_columns if col not in columns]
    if missing:
        messages.error(request, f"Colonnes manquantes : {', '.join(missing)}.")
        return None
    return columns, rows


@login_required
def upload_students(request):
    """Importe des élèves depuis un fichier .xlsx.

    Colonnes attendues : matricule, nom, postnom, prenom, adresse, telephone.
    La position GPS du domicile est définie par le parent à sa première connexion.
    """
    if request.method != "POST":
        return redirect("routing:dashboard")
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked

    parsed = _read_uploaded_excel(request, STUDENT_IMPORT_COLUMNS)
    if parsed is None:
        return redirect("routing:dashboard")
    columns, rows = parsed

    created = updated = skipped = 0
    with recalc_hold():
        for row in rows:
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            matricule = str(row[columns["matricule"]] or "").strip()
            if not matricule:
                skipped += 1
                continue

            defaults = {
                "school": _user_school(request.user),
                "prenom": str(row[columns["prenom"]] or "").strip(),
                "postnom": str(row[columns["postnom"]] or "").strip(),
                "nom": str(row[columns["nom"]] or "").strip(),
                "address": str(row[columns["adresse"]] or "").strip(),
                "parent_phone": str(row[columns["telephone"]] or "").strip(),
            }
            student, was_created = Student.objects.update_or_create(
                matricule=matricule, defaults=defaults
            )
            if was_created:
                created += 1
                # Compte de connexion automatique (mot de passe par défaut = matricule).
                _ensure_student_account(student)
            else:
                updated += 1
    # Les signaux sont suspendus pendant l'import : un seul recalcul global
    # répartit automatiquement tous les élèves importés sur les bus.
    recalculate_school_routes(_user_school(request.user))

    messages.success(
        request,
        f"Import des élèves terminé : {created} créé(s), {updated} mis à jour, "
        f"{skipped} ligne(s) ignorée(s).",
    )
    return redirect("routing:dashboard")


@login_required
def upload_buses(request):
    """Importe des bus depuis un fichier .xlsx.

    Colonnes attendues : code_bus, nom_chauffeur, capacite.
    """
    if request.method != "POST":
        return redirect("routing:dashboard")
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked

    parsed = _read_uploaded_excel(request, BUS_IMPORT_COLUMNS)
    if parsed is None:
        return redirect("routing:dashboard")
    columns, rows = parsed

    created = updated = skipped = 0
    with recalc_hold():
        for row in rows:
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            code_bus = str(row[columns["code_bus"]] or "").strip()
            if not code_bus:
                skipped += 1
                continue
            try:
                capacity = int(float(row[columns["capacite"]]))
            except (TypeError, ValueError):
                skipped += 1
                continue

            _, was_created = Bus.objects.update_or_create(
                code_bus=code_bus,
                defaults={
                    "school": _user_school(request.user),
                    "capacity": capacity,
                    "driver_name": str(row[columns["nom_chauffeur"]] or "").strip(),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
    # Signaux suspendus pendant l'import : un seul recalcul global répartit
    # automatiquement les élèves sur la flotte importée.
    recalculate_school_routes(_user_school(request.user))

    messages.success(
        request,
        f"Import des bus terminé : {created} créé(s), {updated} mis à jour, "
        f"{skipped} ligne(s) ignorée(s).",
    )
    return redirect("routing:dashboard")


def _build_students_excel(school, buses, include_unassigned=True, sheet_title="Élèves par bus"):
    """Construit le classeur Excel des élèves des bus donnés (une ligne par élève)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl import Workbook

    absent = absent_student_ids()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    headers = [
        "Code bus", "Chauffeur", "Capacité", "Matricule",
        "Prénom", "Postnom", "Nom", "Adresse", "Téléphone parent",
        "Statut", "Gelé", "Absent",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="2563EB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    def _student_row(bus, student):
        status = "Pris" if student.is_taken else "Restant"
        return [
            bus.code_bus if bus else "",
            bus.driver_name if bus else "",
            bus.capacity if bus else "",
            student.matricule,
            student.prenom,
            student.postnom,
            student.nom,
            student.address,
            student.parent_phone,
            status,
            "Oui" if student.is_frozen else "Non",
            "Oui" if student.id in absent else "",
        ]

    total = 0
    for bus in buses:
        students = Student.objects.filter(
            school=school, assigned_route__bus=bus, is_active=True
        ).order_by("nom", "prenom")
        for s in students:
            ws.append(_student_row(bus, s))
            total += 1

    if include_unassigned:
        # Élèves actifs sans bus
        unassigned = (
            Student.objects.filter(
                school=school, is_active=True, assigned_route__isnull=True
            ).order_by("nom", "prenom")
        )
        for s in unassigned:
            ws.append(_student_row(None, s))
            total += 1

    if total == 0:
        ws.append(["Aucun élève actif dans cette école."])

    widths = [10, 18, 9, 14, 16, 16, 16, 32, 16, 10, 7, 7]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"
    return wb


@login_required
def export_students_by_bus(request):
    """Export Excel : liste des élèves assignés à chaque bus de l'école.

    Une ligne par élève avec le bus auquel il est rattaché ; les élèves
    sans bus sont listés à la fin (colonne « Code bus » vide).
    """
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    school = _user_school(request.user)
    buses = Bus.objects.filter(school=school).order_by("code_bus")
    wb = _build_students_excel(school, buses)

    filename = re.sub(r"[^\w\-]", "_", school.code_ecole) or "ecole"
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="eleves_par_bus_{filename}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def export_bus_students(request, pk):
    """Export Excel : élèves assignés à UN bus précis de l'école."""
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    school = _user_school(request.user)
    bus = get_object_or_404(Bus, pk=pk)
    if bus.school_id != school.pk:
        messages.error(request, "Ce bus ne fait pas partie de votre école.")
        return redirect("routing:dashboard")
    wb = _build_students_excel(
        school, [bus], include_unassigned=False, sheet_title=bus.code_bus
    )

    filename = re.sub(r"[^\w\-]", "_", bus.code_bus) or "bus"
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="eleves_bus_{filename}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def route_sheet_pdf(request, pk):
    """Export PDF : feuille de route imprimable d'un bus (pour le chauffeur).

    Liste ordonnée des élèves à récupérer (absents du jour exclus), avec
    l'école, le bus, le chauffeur et les indicateurs du trajet.
    """
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked
    school = _user_school(request.user)
    bus = get_object_or_404(Bus, pk=pk)
    if bus.school_id != school.pk:
        messages.error(request, "Ce bus ne fait pas partie de votre école.")
        return redirect("routing:dashboard")

    # Blocage de démarrage : pas de feuille de route tant que des élèves avec
    # position GPS valide ne sont pas affectés à un bus (évite les oublis).
    unassigned = _unassigned_students_with_gps(school)
    if unassigned.exists():
        messages.error(
            request,
            f"Feuille de route bloquée : {unassigned.count()} élève(s) avec position GPS "
            "ne sont pas encore affectés au transport (aucun bus disponible ou capacité "
            "insuffisante). L'affectation est entièrement automatique : elle se fera dès "
            "qu'un bus a de la place.",
        )
        return redirect("routing:dashboard")

    route = bus.routes.order_by("-created_at").first()
    absent = absent_student_ids()
    students = (
        [s for s in _route_students_in_order(route) if s.id not in absent]
        if route
        else []
    )
    taken = sum(1 for s in students if s.is_taken)
    remaining = len(students) - taken

    import reportlab
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Flowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # Police Vera embarquée dans reportlab (Unicode : ✓ / ☐ et accents français).
    vera_dir = os.path.join(os.path.dirname(reportlab.__file__), "fonts")
    try:
        pdfmetrics.registerFont(TTFont("Vera", os.path.join(vera_dir, "Vera.ttf")))
        pdfmetrics.registerFont(TTFont("VeraBd", os.path.join(vera_dir, "VeraBd.ttf")))
    except Exception:
        pass

    INDIGO = colors.HexColor("#2563EB")
    AMBER = colors.HexColor("#F59E0B")
    DARK = colors.HexColor("#1E293B")
    GRAY = colors.HexColor("#64748B")
    WHITE = colors.white

    st_title = ParagraphStyle("title", fontName="VeraBd", fontSize=20, textColor=INDIGO, alignment=TA_LEFT)
    st_sub = ParagraphStyle("sub", fontName="Vera", fontSize=9, textColor=GRAY, alignment=TA_LEFT)
    st_school = ParagraphStyle("school", fontName="VeraBd", fontSize=11, textColor=DARK)
    st_brand = ParagraphStyle("brand", fontName="VeraBd", fontSize=9, textColor=AMBER, alignment=TA_RIGHT)
    st_label = ParagraphStyle("label", fontName="Vera", fontSize=7.5, textColor=GRAY)
    st_value = ParagraphStyle("value", fontName="VeraBd", fontSize=11, textColor=DARK)
    st_cell = ParagraphStyle("cell", fontName="Vera", fontSize=8.5, textColor=DARK)
    st_cell_b = ParagraphStyle("cellb", fontName="VeraBd", fontSize=8.5, textColor=DARK)
    st_hdr = ParagraphStyle("hdr", fontName="VeraBd", fontSize=8.5, textColor=WHITE, alignment=TA_CENTER)

    class _Checkbox(Flowable):
        """Case à cocher vectorielle (rectangle + coche éventuelle)."""

        def __init__(self, checked=False):
            super().__init__()
            self.checked = checked
            self.width = 8
            self.height = 8

        def draw(self):
            self.canv.setLineWidth(0.9)
            self.canv.setStrokeColor(DARK)
            if self.checked:
                self.canv.line(1.8, 4.2, 3.4, 5.9)
                self.canv.line(3.4, 5.9, 6.6, 1.9)
            self.canv.rect(0.6, 0.6, self.width - 1.2, self.height - 1.2, stroke=1, fill=0)

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Vera", 7.5)
        canvas.setFillColor(GRAY)
        canvas.drawString(
            14 * mm, 8 * mm,
            "Document généré par ScolaLoop - Transport Scolaire Intelligent (ESTECH)",
        )
        canvas.drawRightString(A4[0] - 14 * mm, 8 * mm, f"Page {doc.page}")
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=16 * mm,
        title=f"Feuille de route {bus.code_bus}",
        author="ScolaLoop",
    )
    story = []

    # ----- En-tête : école / marque -----
    header = Table(
        [
            [
                Paragraph(school.name, st_school),
                Paragraph("ScolaLoop<br/>Transport Scolaire Intelligent", st_brand),
            ]
        ],
        colWidths=[110 * mm, 72 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW", (0, 0), (-1, 0), 1.2, INDIGO),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(header)
    if school.address:
        story.append(Paragraph(school.address, st_sub))
    story.append(Spacer(1, 6 * mm))

    # ----- Titre -----
    story.append(Paragraph("FEUILLE DE ROUTE", st_title))
    story.append(
        Paragraph(
            f"Bus {bus.code_bus} - {bus.driver_name or 'chauffeur non renseigné'} - "
            f"éditée le {timezone.localdate().strftime('%d/%m/%Y')} à "
            f"{timezone.localtime().strftime('%H:%M')}",
            st_sub,
        )
    )
    story.append(Spacer(1, 5 * mm))

    # ----- Bloc infos bus / trajet -----
    distance = route.total_distance_km if route else 0.0
    duration = route.estimated_duration_minutes if route else 0.0
    info = Table(
        [
            [
                Paragraph("BUS", st_label),
                Paragraph("CHAUFFEUR", st_label),
                Paragraph("CAPACITÉ", st_label),
            ],
            [
                Paragraph(bus.code_bus, st_value),
                Paragraph(bus.driver_name or "-", st_value),
                Paragraph(str(bus.capacity), st_value),
            ],
            [
                Paragraph("DISTANCE", st_label),
                Paragraph("DURÉE ESTIMÉE", st_label),
                Paragraph("ÉLÈVES (PRIS / RESTANTS)", st_label),
            ],
            [
                Paragraph(f"{distance:.2f} km", st_value),
                Paragraph(f"{duration:.0f} min", st_value),
                Paragraph(f"{taken} / {remaining}", st_value),
            ],
        ],
        colWidths=[61 * mm, 61 * mm, 60 * mm],
    )
    info.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#E2E8F0")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(info)
    story.append(Spacer(1, 6 * mm))

    # ----- Tableau des élèves -----
    if students:
        story.append(Paragraph("ÉLÈVES À RÉCUPÉRER", st_title.clone("subtitle", textColor=DARK, fontSize=12)))
        story.append(Spacer(1, 3 * mm))
        head = [Paragraph(h, st_hdr) for h in ["N°", "Matricule", "Nom complet", "Adresse", "Téléphone", "Statut", "Coche"]]
        rows = [head]
        for i, s in enumerate(students, start=1):
            name = " ".join(filter(None, [s.prenom, s.postnom, s.nom]))
            status = Paragraph("PRIS", st_cell_b) if s.is_taken else Paragraph("Restant", st_cell)
            rows.append(
                [
                    Paragraph(str(i), st_cell),
                    Paragraph(s.matricule, st_cell_b),
                    Paragraph(name, st_cell),
                    Paragraph(s.address or "-", st_cell),
                    Paragraph(s.parent_phone or "-", st_cell),
                    status,
                    _Checkbox(checked=s.is_taken),
                ]
            )
        table = Table(rows, colWidths=[9 * mm, 24 * mm, 42 * mm, 56 * mm, 24 * mm, 17 * mm, 10 * mm], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), INDIGO),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, colors.HexColor("#F1F5F9")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 4 * mm))
        story.append(
            Paragraph(
                f"Total : {len(students)} élève(s) - {taken} pris, {remaining} restant(s). "
                "Cochez chaque élève au moment du ramassage.",
                st_sub,
            )
        )
    else:
        story.append(Paragraph("Aucun élève à récupérer.", st_cell))
        if route is None:
            story.append(Paragraph("Aucun trajet n'a encore été généré pour ce bus.", st_sub))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)

    filename = re.sub(r"[^\w\-]", "_", bus.code_bus) or "bus"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="feuille_route_{filename}.pdf"'
    response.write(buf.getvalue())
    return response


@login_required
def upload_students_freeze(request):
    """Gèle un groupe d'élèves depuis un fichier .xlsx dont la PREMIÈRE
    colonne contient obligatoirement les matricules (les autres colonnes
    sont ignorées)."""
    if request.method != "POST":
        return redirect("routing:dashboard")
    blocked = _school_scope_guard(request)
    if blocked:
        return blocked

    file = request.FILES.get("file")
    if file is None:
        messages.error(request, "Aucun fichier sélectionné.")
        return redirect("routing:dashboard")
    if not file.name.lower().endswith(".xlsx"):
        messages.error(request, "Le fichier doit être au format .xlsx.")
        return redirect("routing:dashboard")
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Lecture du fichier Excel impossible : %s", exc)
        messages.error(request, "Impossible de lire le fichier Excel.")
        return redirect("routing:dashboard")

    school = _user_school(request.user)
    rows = workbook.active.iter_rows(values_only=True)
    frozen = not_found = skipped = 0
    first = True
    # Signaux suspendus pendant le gel de groupe : un seul recalcul à la fin.
    with recalc_hold():
        for row in rows:
            if row is None or len(row) == 0:
                continue
            value = row[0]
            if value is None or str(value).strip() == "":
                skipped += 1
                continue
            matricule = str(value).strip()
            if first:
                first = False
                # En-tête éventuel ("Matricule", "matricules"...) : on le saute.
                if matricule.lower() in ("matricule", "matricules", "matricule(s)", "code", "id"):
                    continue
            student = Student.objects.filter(school=school, matricule=matricule).first()
            if student is None:
                not_found += 1
                continue
            if not student.is_frozen:
                student.is_frozen = True
                student.save(update_fields=["is_frozen"])
            frozen += 1
    recalculate_school_routes(school)

    messages.success(
        request,
        f"Gel de groupe terminé : {frozen} élève(s) gelé(s), "
        f"{not_found} matricule(s) introuvable(s), {skipped} ligne(s) vide(s) ignorée(s).",
    )
    return redirect("routing:dashboard")
